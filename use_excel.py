from datetime import datetime
from openpyxl import Workbook


class ExcelUtils(object):
    """
    pip install openpyxl
    pip install pillow
    """

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_two = self.wb.create_sheet('我的表单')
        self.ws_three = self.wb.create_sheet()
        self.ws.title = '你的表单'
        self.ws.sheet_properties.tabColor = 'ff0000'
        self.ws_three = self.wb.create_sheet()


    def do_sth(self):
        self.ws['A1'] = 66
        self.ws['A2'] = '你好'
        self.ws['A3'] = datetime.now()
        self.wb.save('./test.xlsx')


if __name__ == '__main__':
    client = ExcelUtils()
    client.do_sth()
